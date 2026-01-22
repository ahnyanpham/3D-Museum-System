"""
API Routes for Excel Export (Manager only)
Add these routes to your app.py

Requirements:
pip install openpyxl --break-system-packages
"""

from flask import request, jsonify, send_file
from auth import permission_required
import sqlite3
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io

# ============================================================================
# EXCEL EXPORT ROUTES (Manager/Admin only)
# ============================================================================

@app.route('/api/export/customers', methods=['GET'])
@permission_required('export')
def api_export_customers():
    """Export customers to Excel"""
    try:
        db = get_db_connection()
        cursor = db.cursor()
        
        # Get customers data
        cursor.execute("""
            SELECT 
                CUSTOMER_ID, FULLNAME, PHONE, EMAIL, ID_NUMBER,
                NATIONALITY, BIRTH_DATE, GENDER, CREATED_AT
            FROM CUSTOMER
            ORDER BY CREATED_AT DESC
        """)
        
        customers = cursor.fetchall()
        db.close()
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Khách Hàng"
        
        # Define styles
        header_fill = PatternFill(start_color="C41E3A", end_color="C41E3A", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = [
            'ID', 'Họ Tên', 'Số Điện Thoại', 'Email', 'CMND/CCCD',
            'Quốc Tịch', 'Ngày Sinh', 'Giới Tính', 'Ngày Đăng Ký'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Data rows
        for row_num, customer in enumerate(customers, 2):
            for col_num, value in enumerate(customer, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Adjust column widths
        column_widths = [8, 25, 15, 30, 15, 15, 12, 10, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # Add summary
        summary_row = len(customers) + 3
        ws.cell(row=summary_row, column=1, value='Tổng số khách hàng:')
        ws.cell(row=summary_row, column=2, value=len(customers))
        ws.cell(row=summary_row, column=1).font = Font(bold=True)
        
        ws.cell(row=summary_row + 1, column=1, value='Ngày xuất:')
        ws.cell(row=summary_row + 1, column=2, value=datetime.now().strftime('%d/%m/%Y %H:%M:%S'))
        ws.cell(row=summary_row + 1, column=1).font = Font(bold=True)
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f'Khach_Hang_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Lỗi: {str(e)}'
        }), 500


@app.route('/api/export/tickets', methods=['GET'])
@permission_required('export')
def api_export_tickets():
    """Export tickets to Excel"""
    # Get date range from query params
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    status = request.args.get('status', '')
    
    try:
        db = get_db_connection()
        cursor = db.cursor()
        
        # Build query
        query = """
            SELECT 
                t.TICKET_ID, t.TICKET_CODE, tt.TYPE_NAME, c.FULLNAME,
                c.PHONE, t.QUANTITY, t.TOTAL_PRICE, t.PURCHASE_DATE,
                t.STATUS, t.PAYMENT_METHOD,
                vh.CHECK_IN_TIME, vh.CHECK_OUT_TIME, vh.DURATION_MINUTES,
                vh.RATING, vh.FEEDBACK
            FROM TICKET t
            JOIN TICKET_TYPE tt ON t.TICKET_TYPE_ID = tt.TICKET_TYPE_ID
            JOIN CUSTOMER c ON t.CUSTOMER_ID = c.CUSTOMER_ID
            LEFT JOIN VISIT_HISTORY vh ON t.TICKET_ID = vh.TICKET_ID
            WHERE 1=1
        """
        params = []
        
        if start_date:
            query += " AND DATE(t.PURCHASE_DATE) >= ?"
            params.append(start_date)
        
        if end_date:
            query += " AND DATE(t.PURCHASE_DATE) <= ?"
            params.append(end_date)
        
        if status:
            query += " AND t.STATUS = ?"
            params.append(status)
        
        query += " ORDER BY t.PURCHASE_DATE DESC"
        
        cursor.execute(query, params)
        tickets = cursor.fetchall()
        db.close()
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Vé Bán Ra"
        
        # Define styles
        header_fill = PatternFill(start_color="C41E3A", end_color="C41E3A", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Status color mapping
        status_colors = {
            'active': 'C6EFCE',      # Light green
            'checked_in': 'C9DAF8',  # Light blue
            'checked_out': 'FFE599', # Light yellow
            'completed': 'D9D9D9',   # Light gray
            'expired': 'F4CCCC',     # Light red
            'cancelled': 'EA9999'    # Red
        }
        
        # Headers
        headers = [
            'ID Vé', 'Mã Vé', 'Loại Vé', 'Khách Hàng', 'SĐT',
            'Số Lượng', 'Tổng Tiền', 'Ngày Mua', 'Trạng Thái',
            'Thanh Toán', 'Check-in', 'Check-out', 'Thời Gian (phút)',
            'Đánh Giá', 'Phản Hồi'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Data rows
        status_labels = {
            'active': 'Hoạt động',
            'checked_in': 'Check-in',
            'checked_out': 'Check-out',
            'completed': 'Hoàn thành',
            'expired': 'Hết hạn',
            'cancelled': 'Đã hủy'
        }
        
        for row_num, ticket in enumerate(tickets, 2):
            for col_num, value in enumerate(ticket, 1):
                cell = ws.cell(row=row_num, column=col_num)
                
                # Format status column
                if col_num == 9:  # Status column
                    cell.value = status_labels.get(value, value)
                    # Apply status color
                    if value in status_colors:
                        cell.fill = PatternFill(
                            start_color=status_colors[value],
                            end_color=status_colors[value],
                            fill_type="solid"
                        )
                else:
                    cell.value = value
                
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Adjust column widths
        column_widths = [8, 15, 15, 20, 12, 10, 12, 18, 12, 12, 18, 18, 12, 10, 30]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # Add summary
        summary_row = len(tickets) + 3
        
        # Calculate totals
        total_tickets = len(tickets)
        total_revenue = sum(ticket[6] for ticket in tickets if ticket[6])
        
        ws.cell(row=summary_row, column=1, value='Thống kê:')
        ws.cell(row=summary_row, column=1).font = Font(bold=True, size=12)
        
        ws.cell(row=summary_row + 1, column=1, value='Tổng số vé:')
        ws.cell(row=summary_row + 1, column=2, value=total_tickets)
        ws.cell(row=summary_row + 1, column=1).font = Font(bold=True)
        
        ws.cell(row=summary_row + 2, column=1, value='Tổng doanh thu:')
        ws.cell(row=summary_row + 2, column=2, value=f'{total_revenue:,.0f} VNĐ')
        ws.cell(row=summary_row + 2, column=1).font = Font(bold=True)
        ws.cell(row=summary_row + 2, column=2).font = Font(color="C41E3A", bold=True)
        
        # Status breakdown
        if tickets:
            status_counts = {}
            for ticket in tickets:
                status = ticket[8]
                status_counts[status] = status_counts.get(status, 0) + 1
            
            ws.cell(row=summary_row + 4, column=1, value='Chi tiết trạng thái:')
            ws.cell(row=summary_row + 4, column=1).font = Font(bold=True)
            
            row = summary_row + 5
            for status, count in status_counts.items():
                ws.cell(row=row, column=1, value=f'  {status_labels.get(status, status)}:')
                ws.cell(row=row, column=2, value=count)
                row += 1
        
        ws.cell(row=summary_row + 10, column=1, value='Ngày xuất:')
        ws.cell(row=summary_row + 10, column=2, value=datetime.now().strftime('%d/%m/%Y %H:%M:%S'))
        ws.cell(row=summary_row + 10, column=1).font = Font(bold=True)
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f'Ve_Ban_Ra_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Lỗi: {str(e)}'
        }), 500


@app.route('/api/export/revenue', methods=['GET'])
@permission_required('export')
def api_export_revenue_report():
    """Export revenue report to Excel"""
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    
    try:
        db = get_db_connection()
        cursor = db.cursor()
        
        # Get revenue by date
        query = """
            SELECT 
                DATE(PURCHASE_DATE) as sale_date,
                COUNT(*) as ticket_count,
                SUM(TOTAL_PRICE) as daily_revenue
            FROM TICKET
            WHERE 1=1
        """
        params = []
        
        if start_date:
            query += " AND DATE(PURCHASE_DATE) >= ?"
            params.append(start_date)
        
        if end_date:
            query += " AND DATE(PURCHASE_DATE) <= ?"
            params.append(end_date)
        
        query += " GROUP BY DATE(PURCHASE_DATE) ORDER BY sale_date DESC"
        
        cursor.execute(query, params)
        daily_revenue = cursor.fetchall()
        
        # Get revenue by ticket type
        query2 = """
            SELECT 
                tt.TYPE_NAME,
                COUNT(*) as ticket_count,
                SUM(t.TOTAL_PRICE) as type_revenue
            FROM TICKET t
            JOIN TICKET_TYPE tt ON t.TICKET_TYPE_ID = tt.TICKET_TYPE_ID
            WHERE 1=1
        """
        params2 = []
        
        if start_date:
            query2 += " AND DATE(t.PURCHASE_DATE) >= ?"
            params2.append(start_date)
        
        if end_date:
            query2 += " AND DATE(t.PURCHASE_DATE) <= ?"
            params2.append(end_date)
        
        query2 += " GROUP BY tt.TYPE_NAME ORDER BY type_revenue DESC"
        
        cursor.execute(query2, params2)
        type_revenue = cursor.fetchall()
        
        db.close()
        
        # Create Excel workbook
        wb = Workbook()
        
        # Sheet 1: Daily revenue
        ws1 = wb.active
        ws1.title = "Doanh Thu Theo Ngày"
        
        header_fill = PatternFill(start_color="C41E3A", end_color="C41E3A", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers for sheet 1
        headers1 = ['Ngày', 'Số Vé Bán', 'Doanh Thu (VNĐ)']
        for col_num, header in enumerate(headers1, 1):
            cell = ws1.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Data for sheet 1
        for row_num, row_data in enumerate(daily_revenue, 2):
            for col_num, value in enumerate(row_data, 1):
                cell = ws1.cell(row=row_num, column=col_num)
                cell.value = value
                if col_num == 3:  # Revenue column
                    cell.number_format = '#,##0'
                cell.border = border
                cell.alignment = Alignment(horizontal='center' if col_num <= 2 else 'right', vertical='center')
        
        ws1.column_dimensions['A'].width = 15
        ws1.column_dimensions['B'].width = 15
        ws1.column_dimensions['C'].width = 20
        
        # Sheet 2: Revenue by type
        ws2 = wb.create_sheet(title="Doanh Thu Theo Loại Vé")
        
        headers2 = ['Loại Vé', 'Số Vé Bán', 'Doanh Thu (VNĐ)']
        for col_num, header in enumerate(headers2, 1):
            cell = ws2.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        for row_num, row_data in enumerate(type_revenue, 2):
            for col_num, value in enumerate(row_data, 1):
                cell = ws2.cell(row=row_num, column=col_num)
                cell.value = value
                if col_num == 3:
                    cell.number_format = '#,##0'
                cell.border = border
                cell.alignment = Alignment(horizontal='left' if col_num == 1 else 'right', vertical='center')
        
        ws2.column_dimensions['A'].width = 20
        ws2.column_dimensions['B'].width = 15
        ws2.column_dimensions['C'].width = 20
        
        # Add summary to both sheets
        total_revenue = sum(row[2] for row in daily_revenue)
        total_tickets = sum(row[1] for row in daily_revenue)
        
        for ws in [ws1, ws2]:
            summary_row = len(daily_revenue if ws == ws1 else type_revenue) + 3
            ws.cell(row=summary_row, column=1, value='Tổng doanh thu:')
            ws.cell(row=summary_row, column=2, value=f'{total_revenue:,.0f} VNĐ')
            ws.cell(row=summary_row, column=1).font = Font(bold=True)
            ws.cell(row=summary_row, column=2).font = Font(color="C41E3A", bold=True)
            
            ws.cell(row=summary_row + 1, column=1, value='Tổng số vé:')
            ws.cell(row=summary_row + 1, column=2, value=total_tickets)
            ws.cell(row=summary_row + 1, column=1).font = Font(bold=True)
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f'Bao_Cao_Doanh_Thu_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Lỗi: {str(e)}'
        }), 500
